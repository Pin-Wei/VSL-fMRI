import os
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def cbind_Q_type(df):
    return df['Triplet_Pair'] + "_" + df['Question_type']

red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEE99', end_color='FFEE99', fill_type='solid')
green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
greater_than_0 = CellIsRule(operator='greaterThan', formula=['0'], fill=red_fill)

top_dir = '/media/data1/pinwei/SL_hippocampus/behavioral_data'

out_file_1 = os.path.join(top_dir, 'logs', 'jigg_task_performance.xlsx')
out_file_2 = os.path.join(top_dir, 'logs', 'performance_summary.xlsx')
out_file_3 = os.path.join(top_dir, 'familarity_test', 'familarity_test_score.csv')
out_file_4 = os.path.join(top_dir, 'familarity_test', 'tidy_all.xlsx')

for out_file in [out_file_1, out_file_2, out_file_3, out_file_4]:
    if os.path.exists(out_file):
        os.remove(out_file)
        fn = out_file.replace(top_dir, '.')
        print(f"\n### File {fn} exists, deleted successfully. ###")

# sidx = int(sys.argv[1])
# subj_indexs = input("\nInput a list of subject ID (3d): ").split()
subj_indexs = ['002', '003', '004', '005', 
         '006', '007', '008', '009', '010', 
         '011', '012', 
         '101', '102', '104', '105', 
         '106', '107', '108', '109', '110', 
         '111', '112', '113', '114', '115', 
         '116', '117', '118', '119', '120', 
         '121', '122', '123', '124', '125', 
         '126', '127']

for sidx in subj_indexs:
    if sidx[0] == '0':
        prefix = 'PW' # Rapid event-related
    elif sidx[0] == '1':
        prefix = 'Slow'

#1# Jiggle task performance: [mean RT], [STD of RT], number of [Hit]/[miss]/[false alarm (FA)] trials.
    log_file = os.path.join('logs', prefix+sidx, '{:}_jigg_task.xlsx'.format(sidx))
    sheet_list = ['RUN_{:}'.format(run) for run in range(1,11)]
    task_perf = {}

    for run in sheet_list:
        task_perf[run] = {}
        run_data = pd.read_excel(log_file, sheet_name=run, engine='openpyxl')
        rt_stats = run_data.describe().loc[['mean', 'std'], 'RT'].to_dict()
            # Return unbiased standard deviation (i.e., normalized by N-1) by default
        task_perf[run] = {'RT_mean': rt_stats['mean'], 'RT_std': rt_stats['std']}
        task_perf[run]['Hit'] = len(run_data[(run_data.Task == 1) & (run_data.Press == 1)])
        task_perf[run]['miss'] = len(run_data[(run_data.Task == 1) & (run_data.Press == 0)])
        task_perf[run]['FA'] = len(run_data[(run_data.Task == 0) & (run_data.Press == 1)])
    
    M1 = 'w' if not os.path.exists(out_file_1) else 'a'
    with pd.ExcelWriter(out_file_1, engine='openpyxl', mode=M1) as writer:
        pd.DataFrame(task_perf).T.to_excel(writer, sheet_name=str(sidx))

        worksheet = writer.sheets[str(sidx)]
        worksheet.conditional_formatting.add(f'E2:F12', greater_than_0)

#2# Jiggle task performance: summarize in a one sheet and concatenate in another.
    task_perf_summ = {sidx: pd.DataFrame(task_perf).T.loc[:, ['Hit', 'miss', 'FA']].sum().to_dict()}
    task_perf_summ[sidx]['RT_mean2'] = pd.DataFrame(task_perf).T.loc[:, 'RT_mean'].mean()
    df_summ = pd.DataFrame(task_perf_summ).T
    df_summ = df_summ.reindex(columns=['RT_mean2', 'Hit', 'miss', 'FA'])

    df_subj = pd.DataFrame(task_perf).T
    df_subj.insert(0, 'SID', int(sidx))

    if not os.path.exists(out_file_2):
        with pd.ExcelWriter(out_file_2, engine='openpyxl', mode='w') as writer:
            df_summ.to_excel(writer, sheet_name='summ_all')
            df_subj.to_excel(writer, sheet_name='cat_all')
    else:
        with pd.ExcelWriter(out_file_2, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_summ.to_excel(writer, sheet_name='summ_all', header=False, startrow=writer.sheets['summ_all'].max_row)
            worksheet1 = writer.sheets['summ_all']
            worksheet1.conditional_formatting.add(f'D2:E40', greater_than_0)            
            
            df_subj.to_excel(writer, sheet_name='cat_all', header=False, startrow=writer.sheets['cat_all'].max_row)
            worksheet2 = writer.sheets['cat_all']
            smaller_than_0 = CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            greater_than_1 = CellIsRule(operator='greaterThan', formula=['1'], fill=yellow_fill)
            worksheet2.conditional_formatting.add(f'F2:G400', greater_than_0)
            worksheet2.conditional_formatting.add(f'C2:C400', smaller_than_0)
            worksheet2.conditional_formatting.add(f'C2:C400', greater_than_1)

#3# Familiarity test score: [mean Violation degree], [ACC], & [Number of correct ans].
    behav_file = os.path.join('familarity_test', '{:}_test_result.xlsx'.format(sidx))
    behav_data = pd.read_excel(behav_file, engine='openpyxl')
    SL_behav = {sidx: {}}

    SL_behav[sidx]['mean_Vio'] = behav_data['Violation_degree'].mean()
    SL_behav[sidx]['ACC'] = behav_data['correct'].mean()
    SL_behav[sidx]['N_correct'] = sum(behav_data['correct'])

    if not os.path.exists(out_file_3):
        pd.DataFrame(SL_behav).T.to_csv(out_file_3, mode='w')
    else:
        pd.DataFrame(SL_behav).T.to_csv(out_file_3, mode='a', header=False)

#4# Tidy the familiarity test performance.
    behav_data.insert(0, 'question_type', behav_data.apply(cbind_Q_type, axis=1)) # e.g., 'Triplet'+'2-AFC_1'
    tidy_test = dict.fromkeys(behav_data['question_type'], [])
    behav_data = behav_data.sort_values('Triplet_ID')
    for q_type in tidy_test.keys():
        tidy_test[q_type] = list(behav_data.query("question_type == @q_type")["correct"])

    M4 = 'w' if not os.path.exists(out_file_4) else 'a'
    with pd.ExcelWriter(out_file_4, engine='openpyxl', mode=M4) as writer:
        pd.DataFrame(tidy_test).to_excel(writer, sheet_name=str(sidx), index=[1, 2, 3, 4])

        worksheet = writer.sheets[str(sidx)]
        is_Incorrect = CellIsRule(operator='equal', formula=['0'], fill=red_fill)
        is_Correct = CellIsRule(operator='equal', formula=['1'], fill=green_fill)
        worksheet.conditional_formatting.add(f'B2:P5', is_Incorrect)
        worksheet.conditional_formatting.add(f'B2:P5', is_Correct)